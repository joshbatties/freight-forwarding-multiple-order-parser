import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_shipping_template(filename="shipping_order_template.xlsx"):
    # Create a workbook
    wb = Workbook()

    # Create sheets
    instructions_sheet = wb.active
    instructions_sheet.title = "Instructions"
    orders_sheet = wb.create_sheet("Orders")
    reference_sheet = wb.create_sheet("Reference")

    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    
    subheader_fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
    subheader_font = Font(name="Arial", size=11, bold=True)
    
    required_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Create Reference Sheet with lookup values
    reference_sheet["A1"] = "Countries"
    reference_sheet["B1"] = "Hazardous Options"
    
    countries = ["USA", "Canada", "Mexico", "UK", "France", "Germany", "China", "Japan", "Australia", "Brazil", "India"]
    hazardous_options = ["Yes", "No"]
        
    for i, val in enumerate(countries):
        reference_sheet[f"A{i+2}"] = val
        
    for i, val in enumerate(hazardous_options):
        reference_sheet[f"B{i+2}"] = val
    
    # Apply styles to reference sheet
    for col in ['A', 'B']:
        cell = reference_sheet[f"{col}1"]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Create Instructions Sheet
    instructions_sheet["A1"] = "SHIPPING ORDER TEMPLATE INSTRUCTIONS"
    instructions_sheet["A1"].font = Font(name="Arial", size=16, bold=True)
    
    instructions = [
        "",
        "This template is designed to help you submit multiple shipping orders efficiently. Please follow these instructions:",
        "",
        "1. Navigate to the 'Orders' sheet to enter your shipping information.",
        "2. Each row represents one shipping order. You can enter multiple orders in separate rows.",
        "3. Required fields are highlighted in yellow. Please ensure all required fields are completed.",
        "4. Use the dropdown menus where available to select from predefined options.",
        "5. Dates should be entered in the format YYYY-MM-DD (e.g., 2025-04-15).",
        "6. Do not modify the header row or column structure.",
        "7. Do not merge cells as this will affect our processing system.",
        "8. For any questions or assistance, please contact support@shippingcompany.com",
        "",
        "COLUMN DESCRIPTIONS:",
        ""
    ]
    
    for i, text in enumerate(instructions):
        instructions_sheet[f"A{i+2}"] = text
        
    # Define field descriptions
    field_descriptions = [
        ("Customer Information", ""),
        ("customer_code", "Your unique customer code provided by our company"),
        ("primary_contact", "Name of the person responsible for this shipment"),
        ("contact_email", "Email address for shipping notifications"),
        ("contact_phone", "Phone number for urgent communications"),
        ("", ""),
        ("Order Details", ""),
        ("po_number", "Your Purchase Order number for this shipment"),
        ("pickup_date", "Requested date for collection (YYYY-MM-DD)"),
        ("delivery_date", "Requested delivery date (YYYY-MM-DD)"),
        ("", ""),
        ("Shipment Details", ""),
        ("hs_code", "Harmonized System code for the goods"),
        ("goods_description", "Description of items being shipped"),
        ("quantity", "Number of items/packages in this shipment"),
        ("weight_kg", "Total weight in kilograms"),
        ("length_cm", "Length of package in centimeters"),
        ("width_cm", "Width of package in centimeters"),
        ("height_cm", "Height of package in centimeters"),
        ("hazardous", "Select 'Yes' if shipment contains hazardous materials"),
        ("declared_value", "Monetary value of shipment in USD for insurance purposes"),
        ("", ""),
        ("Origin Information", ""),
        ("origin_address", "Street address for pickup"),
        ("origin_city", "City for pickup"),
        ("origin_state", "State/Province for pickup"),
        ("origin_country", "Country for pickup (select from dropdown)"),
        ("origin_postal_code", "ZIP/Postal code for pickup location"),
        ("origin_contact", "Contact person at pickup location"),
        ("origin_phone", "Phone number at pickup location"),
        ("", ""),
        ("Destination Information", ""),
        ("destination_address", "Street address for delivery"),
        ("destination_city", "City for delivery"),
        ("destination_state", "State/Province for delivery"),
        ("destination_country", "Country for delivery (select from dropdown)"),
        ("destination_postal_code", "ZIP/Postal code for delivery location"),
        ("destination_contact", "Contact person at delivery location"),
        ("destination_phone", "Phone number at delivery location"),
        ("", ""),
        ("Additional Information", ""),
        ("special_instructions", "Any special requirements or notes for this shipment")
    ]
    
    row = 18
    for field, description in field_descriptions:
        if field and not description:  # This is a section header
            instructions_sheet[f"A{row}"] = field
            instructions_sheet[f"A{row}"].font = subheader_font
            instructions_sheet[f"A{row}"].fill = subheader_fill
        else:
            instructions_sheet[f"A{row}"] = field
            instructions_sheet[f"B{row}"] = description
        row += 1
    
    # Auto-adjust column width
    for col in ['A', 'B']:
        instructions_sheet.column_dimensions[col].width = 25
    
    # Create Orders Sheet
    # Define columns and their properties
    columns = [
        # Customer Information
        {"name": "customer_code", "required": True, "validation": None, "width": 15},
        {"name": "primary_contact", "required": True, "validation": None, "width": 20},
        {"name": "contact_email", "required": True, "validation": None, "width": 25},
        {"name": "contact_phone", "required": True, "validation": None, "width": 15},
        
        # Order Details
        {"name": "po_number", "required": True, "validation": None, "width": 15},
        {"name": "pickup_date", "required": True, "validation": "date", "width": 15},
        {"name": "delivery_date", "required": True, "validation": "date", "width": 15},
        
        # Shipment Details
        {"name": "hs_code", "required": True, "validation": None, "width": 15},
        {"name": "goods_description", "required": True, "validation": None, "width": 30},
        {"name": "quantity", "required": True, "validation": "number", "width": 10},
        {"name": "weight_kg", "required": True, "validation": "number", "width": 10},
        {"name": "length_cm", "required": True, "validation": "number", "width": 10},
        {"name": "width_cm", "required": True, "validation": "number", "width": 10},
        {"name": "height_cm", "required": True, "validation": "number", "width": 10},
        {"name": "hazardous", "required": True, "validation": "hazardous", "width": 10},
        {"name": "declared_value", "required": True, "validation": "number", "width": 15},
        
        # Origin Information
        {"name": "origin_address", "required": True, "validation": None, "width": 30},
        {"name": "origin_city", "required": True, "validation": None, "width": 15},
        {"name": "origin_state", "required": True, "validation": None, "width": 15},
        {"name": "origin_country", "required": True, "validation": "country", "width": 15},
        {"name": "origin_postal_code", "required": True, "validation": None, "width": 15},
        {"name": "origin_contact", "required": True, "validation": None, "width": 20},
        {"name": "origin_phone", "required": True, "validation": None, "width": 15},
        
        # Destination Information
        {"name": "destination_address", "required": True, "validation": None, "width": 30},
        {"name": "destination_city", "required": True, "validation": None, "width": 15},
        {"name": "destination_state", "required": True, "validation": None, "width": 15},
        {"name": "destination_country", "required": True, "validation": "country", "width": 15},
        {"name": "destination_postal_code", "required": True, "validation": None, "width": 15},
        {"name": "destination_contact", "required": True, "validation": None, "width": 20},
        {"name": "destination_phone", "required": True, "validation": None, "width": 15},
        
        # Special Instructions (moved to end)
        {"name": "special_instructions", "required": False, "validation": None, "width": 30}
    ]
    
    # Create section headers and column headers
    headers = []
    section_ranges = {
        "Customer Information": [],
        "Order Details": [],
        "Shipment Details": [],
        "Origin Information": [],
        "Destination Information": [],
        "Additional Information": []
    }
    
    col_index = 1
    for col in columns:
        headers.append(col["name"])
        
        # Track section ranges for formatting
        if col["name"] in ["customer_code", "primary_contact", "contact_email", "contact_phone"]:
            section_ranges["Customer Information"].append(col_index)
        elif col["name"] in ["po_number", "pickup_date", "delivery_date"]:
            section_ranges["Order Details"].append(col_index)
        elif col["name"] in ["hs_code", "goods_description", "quantity", "weight_kg", "length_cm", "width_cm", "height_cm", 
                           "hazardous", "declared_value"]:
            section_ranges["Shipment Details"].append(col_index)
        elif col["name"].startswith("origin"):
            section_ranges["Origin Information"].append(col_index)
        elif col["name"].startswith("destination"):
            section_ranges["Destination Information"].append(col_index)
        elif col["name"] == "special_instructions":
            section_ranges["Additional Information"].append(col_index)
            
        col_index += 1
    
    # Format column headers to be more human-readable
    formatted_headers = []
    for header in headers:
        # Convert snake_case to Title Case with spaces
        words = header.split('_')
        formatted_header = ' '.join(word.capitalize() for word in words)
        formatted_headers.append(formatted_header)
    
    # Add column headers to the orders sheet
    for i, header in enumerate(formatted_headers, 1):
        cell = orders_sheet.cell(row=2, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        
        # Set column width
        orders_sheet.column_dimensions[get_column_letter(i)].width = columns[i-1]["width"]
        
        # Add fill color for required fields
        if columns[i-1]["required"]:
            for row_idx in range(3, 103):  # Apply to rows 3-102 (100 order rows)
                cell = orders_sheet.cell(row=row_idx, column=i)
                cell.fill = required_fill
    
    # Add section headers
    section_start_cols = {}
    section_end_cols = {}
    
    for section, cols in section_ranges.items():
        if cols:  # Only process sections with columns
            section_start_cols[section] = min(cols)
            section_end_cols[section] = max(cols)
    
    row_idx = 1
    for section, start_col in section_start_cols.items():
        end_col = section_end_cols[section]
        
        # First set the value on the cell
        cell = orders_sheet.cell(row=row_idx, column=start_col, value=section)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = Alignment(horizontal="center")
        
        # Then merge the cells if needed
        if start_col != end_col:
            orders_sheet.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
    
    # Add data validations
    # Country validation
    origin_country_col = [i+1 for i, col in enumerate(columns) if col["name"] == "origin_country"][0]
    origin_country_dv = DataValidation(type="list", formula1=f"=Reference!$A$2:$A${len(countries)+1}", allow_blank=False)
    orders_sheet.add_data_validation(origin_country_dv)
    origin_country_dv.add(f"{get_column_letter(origin_country_col)}3:{get_column_letter(origin_country_col)}102")
    
    dest_country_col = [i+1 for i, col in enumerate(columns) if col["name"] == "destination_country"][0]
    dest_country_dv = DataValidation(type="list", formula1=f"=Reference!$A$2:$A${len(countries)+1}", allow_blank=False)
    orders_sheet.add_data_validation(dest_country_dv)
    dest_country_dv.add(f"{get_column_letter(dest_country_col)}3:{get_column_letter(dest_country_col)}102")
    
    # Hazardous validation
    hazardous_col = [i+1 for i, col in enumerate(columns) if col["name"] == "hazardous"][0]
    hazardous_dv = DataValidation(type="list", formula1=f"=Reference!$B$2:$B${len(hazardous_options)+1}", allow_blank=False)
    orders_sheet.add_data_validation(hazardous_dv)
    hazardous_dv.add(f"{get_column_letter(hazardous_col)}3:{get_column_letter(hazardous_col)}102")
    
    # Number validation for numeric fields
    for i, col in enumerate(columns):
        if col["validation"] == "number":
            number_dv = DataValidation(type="decimal", operator="greaterThan", formula1="0")
            orders_sheet.add_data_validation(number_dv)
            number_dv.add(f"{get_column_letter(i+1)}3:{get_column_letter(i+1)}102")
    
    # Date validation
    for i, col in enumerate(columns):
        if col["validation"] == "date":
            cell_range = f"{get_column_letter(i+1)}3:{get_column_letter(i+1)}102"
            orders_sheet.add_data_validation(DataValidation(type="date"))
            
            # Set date format
            for row_idx in range(3, 103):
                cell = orders_sheet.cell(row=row_idx, column=i+1)
                cell.number_format = "YYYY-MM-DD"
    
    # Create an example row in the Instructions sheet instead
    example_row_header = "EXAMPLE ORDER:"
    row = instructions_sheet.max_row + 2  # Add two rows after last content
    instructions_sheet.cell(row=row, column=1, value=example_row_header).font = subheader_font
    
    example_data = {
        "Customer Code": "CUST12345",
        "Primary Contact": "John Smith",
        "Contact Email": "john.smith@abcshipping.com",
        "Contact Phone": "+1-555-123-4567",
        "PO Number": "PO78901",
        "Pickup Date": "2025-04-05",
        "Delivery Date": "2025-04-12",
        "HS Code": "8471.30",
        "Goods Description": "Electronic components - laptop parts",
        "Quantity": "5",
        "Weight Kg": "75",
        "Length Cm": "120",
        "Width Cm": "80",
        "Height Cm": "60",
        "Hazardous": "No",
        "Declared Value": "5000 USD",
        "Origin Address": "123 Industrial Parkway",
        "Origin City": "Boston",
        "Origin State": "MA",
        "Origin Country": "USA",
        "Origin Postal Code": "02110",
        "Origin Contact": "Sarah Johnson",
        "Origin Phone": "+1-555-987-6543",
        "Destination Address": "456 Commerce Street",
        "Destination City": "Los Angeles",
        "Destination State": "CA",
        "Destination Country": "USA",
        "Destination Postal Code": "90001",
        "Destination Contact": "Michael Brown",
        "Destination Phone": "+1-555-789-0123",
        "Special Instructions": "Please handle with care. Call recipient before delivery."
    }
    
    row += 1
    for key, value in example_data.items():
        instructions_sheet.cell(row=row, column=1, value=key).font = Font(bold=True)
        instructions_sheet.cell(row=row, column=2, value=value)
        row += 1
    
    # Convert the range to an Excel Table
    # Calculate table range (starts at A2 and extends to the last column and row 102)
    last_column_letter = get_column_letter(len(columns))
    table_range = f"A2:{last_column_letter}102"
    
    # Create the table
    table = Table(displayName="ShippingOrdersTable", ref=table_range)
    
    # Add a default style to the table
    style = TableStyleInfo(
        name="TableStyleMedium2", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    
    # Add the table to the worksheet
    orders_sheet.add_table(table)
    
    # Freeze header rows - freeze at row 3 to keep headers visible when scrolling
    orders_sheet.freeze_panes = "A3"
    
    # Set title and page setup
    orders_sheet.oddHeader.center.text = "Shipping Order Template"
    orders_sheet.oddFooter.center.text = "Page &[Page] of &[Pages]"
    
    # Save the workbook
    wb.save(filename)
    print(f"Template created successfully as {filename}")
    return filename

if __name__ == "__main__":
    create_shipping_template()